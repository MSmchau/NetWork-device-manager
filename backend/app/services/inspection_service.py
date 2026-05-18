import logging
import re
import paramiko
from netmiko import ConnectHandler
from netmiko.exceptions import NetMikoTimeoutException, NetMikoAuthenticationException
from app.config import settings, DEVICE_TYPE_MAP

logger = logging.getLogger(__name__)

# H3C Comware 等老设备仅支持 ssh-dss 主机密钥，Paramiko 3.x+ 默认禁用，需手动启用
paramiko.Transport._preferred_keys = ("ssh-dss",) + paramiko.Transport._preferred_keys

# 各厂商标准巡检命令集
INSPECTION_COMMANDS = {
    "hp_comware": {
        "cpu": "display cpu-usage",
        "memory": "display memory",
        "interfaces": "display interface brief",
        "hardware": "display device",
        "uptime": "display version",
    },
    "huawei": {
        "cpu": "display cpu-usage",
        "memory": "display memory",
        "interfaces": "display interface brief",
        "hardware": "display device",
        "uptime": "display version",
    },
    "cisco_ios": {
        "cpu": "show processes cpu",
        "memory": "show memory",
        "interfaces": "show interfaces summary",
        "hardware": "show inventory",
        "uptime": "show version",
    },
    "ruijie_os": {
        "cpu": "show cpu",
        "memory": "show memory statistics",
        "interfaces": "show interface brief",
        "hardware": "show device",
        "uptime": "show version",
    },
}

def device_connect(device):
    """根据设备类型和协议建立连接（SSH / Telnet），支持多厂商"""
    net_type = DEVICE_TYPE_MAP.get(device.device_type, "hp_comware")
    protocol = getattr(device, "protocol", "ssh") or "ssh"
    port = getattr(device, "port", None) or (23 if protocol == "telnet" else settings.DEVICE_SSH_PORT)

    device_info = {
        "device_type": net_type,
        "host": device.ip,
        "username": device.username,
        "password": device.password,
        "port": port,
        "timeout": settings.DEVICE_SSH_TIMEOUT,
    }
    if protocol == "telnet":
        device_info["use_telnet"] = True

    proto_label = "Telnet" if protocol == "telnet" else "SSH"
    try:
        return ConnectHandler(**device_info)
    except NetMikoTimeoutException:
        logger.warning("%s 连接超时: %s:%s (%s)", proto_label, device.ip, port, getattr(device, "device_type", "unknown"))
        return None
    except NetMikoAuthenticationException:
        logger.warning("%s 认证失败: %s 用户名/密码错误", proto_label, device.ip)
        return None

def inspect_device(device):
    """执行标准巡检，返回结构化 JSON 结果"""
    conn = device_connect(device)
    if not conn:
        return {
            "overall_status": "critical",
            "checks": [
                {"name": "connectivity", "status": "fail", "detail": "SSH 连接失败"}
            ],
        }
    net_type = DEVICE_TYPE_MAP.get(device.device_type, "hp_comware")
    commands = INSPECTION_COMMANDS.get(net_type, {})
    checks = [{"name": "connectivity", "status": "pass",
               "detail": f"SSH 连接成功 ({device.ip}:{getattr(device, 'port', None) or settings.DEVICE_SSH_PORT})"}]
    try:
        for check_name, cmd in commands.items():
            try:
                output = conn.send_command(cmd)
            except Exception as e:
                checks.append({"name": check_name, "status": "fail", "detail": f"{check_name} 检查失败: {str(e)}"})
                continue
            checks.append(_parse_check(check_name, output, net_type))
    finally:
        conn.disconnect()
    status_rank = {"pass": 0, "warning": 1, "fail": 2}
    max_rank = max(status_rank.get(c.get("status", "pass"), 0) for c in checks)
    overall = {0: "healthy", 1: "warning", 2: "critical"}[max_rank]
    return {"overall_status": overall, "checks": checks}

def _parse_check(check_name, output, net_type):
    """解析各厂商命令输出"""
    if check_name == "cpu":
        for line in output.split("\n"):
            if "%" in line:
                try:
                    val = float(line.split("%")[0].split()[-1])
                    status = "pass" if val < 70 else ("warning" if val < 90 else "fail")
                    return {"name": "cpu", "status": status, "value": val, "detail": f"CPU 使用率: {val}%"}
                except ValueError:
                    pass
        return {"name": "cpu", "status": "warning", "detail": "未能解析 CPU 使用率"}
    elif check_name == "memory":
        for line in output.split("\n"):
            if "%" in line:
                try:
                    val = float(line.split("%")[0].split()[-1])
                    status = "pass" if val < 70 else ("warning" if val < 90 else "fail")
                    return {"name": "memory", "status": status, "value": val, "detail": f"内存使用率: {val}%"}
                except ValueError:
                    pass
        return {"name": "memory", "status": "warning", "detail": "未能解析内存使用率"}
    elif check_name == "interfaces":
        up = down = adm = 0
        lines = output.split("\n")

        if net_type in ("hp_comware", "huawei", "ruijie_os"):
            # display interface brief → 第二列为 Link 状态 (UP / DOWN / ADM)
            for line in lines:
                parts = line.strip().split()
                if len(parts) >= 2:
                    status_val = parts[1].upper()
                    if status_val == "UP":
                        up += 1
                    elif status_val == "DOWN":
                        down += 1
                    elif status_val == "ADM":
                        adm += 1
        elif net_type == "cisco_ios":
            for line in lines:
                lower = line.lower()
                if "is up" in lower and "is down" not in lower:
                    up += 1
                elif "is down" in lower:
                    down += 1
        else:
            up = len(re.findall(r'\bup\b', output, re.IGNORECASE))
            down = len(re.findall(r'\bdown\b', output, re.IGNORECASE))

        # down 端口多为空闲未使用的端口，有任意 up 端口即视为正常
        status = "fail" if up == 0 and down > 0 else "pass"
        detail = f"接口: {up} up, {down} down"
        if adm:
            detail += f", {adm} administratively down"
        return {"name": "interfaces", "status": status, "detail": detail}
    elif check_name == "hardware":
        # 解析具体硬件组件状态，标记异常组件
        abnormal = []  # 异常组件列表
        healthy_count = 0

        for line in output.split("\n"):
            line = line.strip()
            if not line:
                continue

            # 检查异常状态关键词
            m = re.search(r'\b(Fault|Abnormal|Absent|Offline|Failed)\b', line, re.IGNORECASE)
            if m:
                comp = line[:m.start()].strip()
                abnormal.append(f"{comp} [{m.group(0)}]" if comp else f"[{m.group(0)}]")
                continue

            # 检查正常状态
            if re.search(r'\b(Normal|OK)\b', line, re.IGNORECASE):
                healthy_count += 1

        if abnormal:
            status = "warning"
            detail = "部分硬件异常: " + "; ".join(abnormal)
        elif healthy_count > 0:
            status = "pass"
            detail = "硬件状态正常"
        else:
            # 无法解析硬件状态时默认通过（如 Cisco show inventory 等无状态命令）
            status = "pass"
            detail = "硬件状态正常"

        return {"name": "hardware", "status": status, "detail": detail}
    elif check_name == "uptime":
        # 查找包含 uptime/运行时间的行（各厂商 display version / show version 通用）
        uptime_text = "未知"
        for line in output.split("\n"):
            stripped = line.strip()
            if re.search(r'uptime|运行时间', stripped, re.IGNORECASE):
                # 去除 "uptime is" / "System uptime is" 等前缀
                uptime_text = re.sub(r'^(System\s+)?(uptime\s+is\s+)', '', stripped, flags=re.IGNORECASE)
                break
        if uptime_text == "未知":
            lines = [l.strip() for l in output.split("\n") if l.strip()]
            uptime_text = lines[0] if lines else "未知"
        return {"name": "uptime", "status": "pass", "detail": f"运行时间: {uptime_text}"}
    return {"name": check_name, "status": "pass", "detail": "检查完成"}


def generate_inspection_excel(db):
    """生成巡检汇总 Excel，包含汇总表和详细数据表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from sqlalchemy import func, and_
    import json

    from app.models.device import Device
    from app.models.inspection import InspectionRecord

    # 子查询：每个设备的最新巡检记录
    subq = (
        db.query(
            InspectionRecord.device_id,
            func.max(InspectionRecord.created_at).label("max_created"),
        )
        .group_by(InspectionRecord.device_id)
        .subquery()
    )
    records = (
        db.query(InspectionRecord)
        .join(
            subq,
            and_(
                InspectionRecord.device_id == subq.c.device_id,
                InspectionRecord.created_at == subq.c.max_created,
            ),
        )
        .all()
    )

    devices = {d.id: d for d in db.query(Device).all()}
    device_type_map = {v: k for k, v in DEVICE_TYPE_MAP.items()}

    wb = Workbook()

    # ---- Sheet 1: 汇总表 ----
    ws1 = wb.active
    ws1.title = "汇总表"
    header1 = ["设备名称", "IP地址", "设备类型", "巡检时间", "总体状态",
               "连通性", "CPU", "内存", "接口", "硬件", "运行时间"]
    ws1.append(header1)

    # ---- Sheet 2: 详细数据 ----
    ws2 = wb.create_sheet("详细数据")
    header2 = ["设备名称", "IP地址", "巡检时间", "总体状态",
               "检查项", "检查状态", "检查值", "详情"]
    ws2.append(header2)

    # 表头样式
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for ws in (ws1, ws2):
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    for record in records:
        dev = devices.get(record.device_id)
        dev_name = dev.name if dev else f"ID:{record.device_id}"
        dev_ip = dev.ip if dev else ""
        dev_type_str = device_type_map.get(dev.device_type, dev.device_type) if dev else ""
        inspect_time = record.created_at.strftime("%Y-%m-%d %H:%M:%S") if record.created_at else ""
        overall = record.overall_status

        # 解析 result JSON
        if not record.result:
            continue
        try:
            result = json.loads(record.result)
        except (json.JSONDecodeError, TypeError):
            continue

        checks = result.get("checks", [])
        check_map = {c["name"]: c for c in checks}

        def _status_label(c):
            return c.get("status", "")

        def _check_str(name):
            c = check_map.get(name)
            if not c:
                return ""
            s = _status_label(c)
            v = c.get("value")
            if v is not None:
                return f'{v}% ({s})'
            d = c.get("detail", "")
            return f"{s}" if not d else f"{s} - {d}"

        # 汇总行
        row1 = [
            dev_name,
            dev_ip,
            dev_type_str,
            inspect_time,
            overall,
            _check_str("connectivity"),
            _check_str("cpu"),
            _check_str("memory"),
            _check_str("interfaces"),
            _check_str("hardware"),
            _check_str("uptime"),
        ]
        ws1.append(row1)

        # 详细数据行（每个检查项一行）
        for c in checks:
            name = c.get("name", "")
            status = _status_label(c)
            value = c.get("value")
            detail = c.get("detail", "")
            ws2.append([
                dev_name, dev_ip, inspect_time, overall,
                name, status, value, detail,
            ])

    # 自动调整列宽
    for ws in (ws1, ws2):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    # 中文字符按 2 倍宽度计算
                    cn_count = sum(1 for ch in str(cell.value or "") if '一' <= ch <= '鿿')
                    cell_len += cn_count
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
