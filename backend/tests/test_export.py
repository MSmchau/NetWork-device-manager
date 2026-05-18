"""测试巡检报告 Excel 导出功能"""
import json
from unittest.mock import MagicMock
from io import BytesIO
from datetime import datetime

from openpyxl import load_workbook

from app.services.inspection_service import generate_inspection_excel


def _make_mock_db(records, devices):
    """构造一个模拟的 db session，返回给定的 records 和 devices"""
    mock_db = MagicMock()
    call_log = {"count": 0}

    def query_side_effect(*_args, **_kwargs):
        call_log["count"] += 1
        q = MagicMock()
        q.group_by.return_value = q
        q.subquery.return_value = MagicMock(name="subq")
        q.join.return_value = q
        q.order_by.return_value = q
        q.offset.return_value = q
        q.limit.return_value = q
        q.count.return_value = len(records)

        # 第一次 query → 子查询（返回空列表）
        # 第二次 query → InspectionRecord（返回 records）
        # 第三次 query → Device（返回 devices）
        if call_log["count"] == 2:
            q.all.return_value = records
        elif call_log["count"] == 3:
            q.all.return_value = devices
        else:
            q.all.return_value = []
        return q

    mock_db.query = query_side_effect
    return mock_db


class MockDevice:
    def __init__(self, id=1, name="test-device", ip="10.0.0.1", device_type="H3C"):
        self.id = id
        self.name = name
        self.ip = ip
        self.device_type = device_type


class MockInspectionRecord:
    def __init__(self, device_id=1, created_at=None, overall_status="healthy", result=None):
        self.device_id = device_id
        self.created_at = created_at or datetime.now()
        self.overall_status = overall_status
        self.result = result


def test_generate_excel_empty():
    """无记录时应生成含表头的空 Excel"""
    mock_db = _make_mock_db([], [])
    result = generate_inspection_excel(mock_db)
    wb = load_workbook(BytesIO(result))
    assert "汇总表" in wb.sheetnames
    assert "详细数据" in wb.sheetnames
    # 只有表头，没有数据行
    assert wb["汇总表"].max_row == 1
    assert wb["详细数据"].max_row == 1
    # 验证表头
    assert wb["汇总表"].cell(1, 1).value == "设备名称"
    assert wb["详细数据"].cell(1, 1).value == "设备名称"


def test_generate_excel_with_data():
    """有巡检记录时应生成包含正确数据的 Excel"""
    records = [
        MockInspectionRecord(
            device_id=1,
            created_at=datetime(2026, 5, 18, 10, 0, 0),
            overall_status="healthy",
            result=json.dumps({
                "overall_status": "healthy",
                "checks": [
                    {"name": "connectivity", "status": "pass", "detail": "SSH 连接成功"},
                    {"name": "cpu", "status": "pass", "value": 45.0, "detail": "CPU 使用率: 45.0%"},
                    {"name": "memory", "status": "warning", "value": 75.0, "detail": "内存使用率: 75.0%"},
                    {"name": "interfaces", "status": "pass", "detail": "接口: 24 up, 3 down"},
                    {"name": "hardware", "status": "pass", "detail": "硬件状态正常"},
                    {"name": "uptime", "status": "pass", "detail": "运行时间: 2 weeks"},
                ],
            }),
        ),
    ]
    devices = [MockDevice(id=1, name="core-sw-01", ip="10.0.0.1", device_type="H3C")]

    mock_db = _make_mock_db(records, devices)
    result = generate_inspection_excel(mock_db)
    wb = load_workbook(BytesIO(result))

    # 汇总表：1 行数据
    ws1 = wb["汇总表"]
    assert ws1.max_row == 2  # 表头 + 1 行数据
    assert ws1.cell(2, 1).value == "core-sw-01"  # 设备名称
    assert ws1.cell(2, 2).value == "10.0.0.1"     # IP
    assert ws1.cell(2, 3).value == "H3C"           # 设备类型
    assert ws1.cell(2, 5).value == "healthy"       # 总体状态

    # 详细数据：每个检查项一行（6 行）
    ws2 = wb["详细数据"]
    assert ws2.max_row == 7  # 表头 + 6 个检查项
    # 检查是否有 cpu 和 memory 的详细数据
    rows = [(ws2.cell(r, 5).value, ws2.cell(r, 6).value) for r in range(2, 8)]
    assert ("cpu", "pass") in rows
    assert ("memory", "warning") in rows


def test_generate_excel_invalid_json():
    """result 字段包含非法 JSON 时应跳过该记录"""
    records = [
        MockInspectionRecord(
            device_id=1,
            overall_status="healthy",
            result="这不是合法的 JSON",
        ),
    ]
    devices = [MockDevice(id=1)]
    mock_db = _make_mock_db(records, devices)
    result = generate_inspection_excel(mock_db)
    wb = load_workbook(BytesIO(result))
    # 非法 JSON 的记录被跳过，只有表头
    assert wb["汇总表"].max_row == 1
